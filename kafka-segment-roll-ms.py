#!/usr/bin/env python3
"""
Kafka Broker - Rolled New Log Segment Analyzer
Sorts partitions by highest to lowest segment roll time.
"""

import re
import sys
import os
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

LINE_PATTERN = re.compile(
    r"\[(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2},\d+)\]"
    r"\s+INFO\s+"
    r"\[MergedLog\s+partition=([\w\.\-]+?),"
    r"\s*dir=([^\]]+)\]"
    r"\s+Rolled new log segment at offset\s+(\d+)"
    r"\s+in\s+(\d+)\s+ms\."
)

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
DATA_FONT = Font(name="Arial", size=10)
NUM_FONT = Font(name="Arial", size=10)
WARN_FILL = PatternFill("solid", fgColor="FFC7CE")
WARN_FONT = Font(name="Arial", size=10, color="9C0006", bold=True)
MOD_FILL = PatternFill("solid", fgColor="FFEB9C")
MOD_FONT = Font(name="Arial", size=10, color="9C6500")
OK_FILL = PatternFill("solid", fgColor="C6EFCE")
OK_FONT = Font(name="Arial", size=10, color="006100")
ALT_FILL = PatternFill("solid", fgColor="D6E4F0")
BORDER = Border(
    bottom=Side(style="thin", color="D9E2F3"),
    top=Side(style="thin", color="D9E2F3"),
    left=Side(style="thin", color="D9E2F3"),
    right=Side(style="thin", color="D9E2F3"),
)
TITLE_FONT = Font(name="Arial", bold=True, size=14, color="1F3864")
SUMMARY_FILL = PatternFill("solid", fgColor="FFF2CC")


def parse_log(filepath):
    entries = []
    with open(filepath, "r", errors="ignore") as f:
        for line in f:
            if "Rolled new log segment at" not in line:
                continue
            m = LINE_PATTERN.search(line)
            if m:
                ts_str, partition, directory, offset, duration_ms = m.groups()
                ts = datetime.strptime(ts_str, "%Y-%m-%d %H:%M:%S,%f")
                entries.append({
                    "timestamp": ts,
                    "date": ts.strftime("%Y-%m-%d"),
                    "time": ts.strftime("%H:%M:%S"),
                    "partition": partition,
                    "directory": directory.strip(),
                    "offset": int(offset),
                    "duration_ms": int(duration_ms),
                })
    return entries


def style_status(cell, duration_ms):
    cell.alignment = Alignment(horizontal="center")
    if duration_ms >= 1000:
        cell.value = "SLOW"
        cell.font = WARN_FONT
        cell.fill = WARN_FILL
    elif duration_ms >= 100:
        cell.value = "MODERATE"
        cell.font = MOD_FONT
        cell.fill = MOD_FILL
    else:
        cell.value = "OK"
        cell.font = OK_FONT
        cell.fill = OK_FILL


def style_duration(cell, duration_ms):
    cell.number_format = "#,##0"
    if duration_ms >= 1000:
        cell.font = WARN_FONT
        cell.fill = WARN_FILL
    elif duration_ms >= 100:
        cell.font = MOD_FONT
        cell.fill = MOD_FILL
    else:
        cell.font = NUM_FONT


def add_headers(ws, row, headers):
    for col_idx, (h, w) in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[row].height = 25


def create_excel(entries, output_path):
    wb = Workbook()

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # SHEET 1: All Entries — Sorted by Duration DESC
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    ws = wb.active
    ws.title = "Rolled Segments (By Duration)"
    ws.sheet_properties.tabColor = "2F5496"

    ws.merge_cells("A1:H1")
    ws["A1"] = "Kafka — Rolled New Log Segment (Sorted: Highest to Lowest Duration)"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A2:H2")
    ws["A2"] = f"Total: {len(entries)} entries | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A2"].font = Font(name="Arial", size=10, italic=True, color="666666")

    headers = [
        ("Rank", 7),
        ("Partition Name", 65),
        ("Duration (ms)", 15),
        ("Duration (sec)", 14),
        ("Offset", 18),
        ("Date", 12),
        ("Time", 10),
        ("Status", 12),
    ]
    add_headers(ws, 4, headers)

    sorted_entries = sorted(entries, key=lambda e: e["duration_ms"], reverse=True)

    for i, e in enumerate(sorted_entries):
        row = i + 5
        ws.cell(row=row, column=1, value=i + 1).font = Font(name="Arial", size=10, bold=True)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=2, value=e["partition"]).font = DATA_FONT
        dur_cell = ws.cell(row=row, column=3, value=e["duration_ms"])
        style_duration(dur_cell, e["duration_ms"])
        ws.cell(row=row, column=4, value=f"=C{row}/1000").font = NUM_FONT
        ws.cell(row=row, column=4).number_format = "0.000"
        ws.cell(row=row, column=5, value=e["offset"]).font = NUM_FONT
        ws.cell(row=row, column=5).number_format = "#,##0"
        ws.cell(row=row, column=6, value=e["date"]).font = DATA_FONT
        ws.cell(row=row, column=7, value=e["time"]).font = DATA_FONT
        style_status(ws.cell(row=row, column=8), e["duration_ms"])

        if i % 2 == 1:
            for col in [1, 2, 5, 6, 7]:
                ws.cell(row=row, column=col).fill = ALT_FILL
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = BORDER

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:H{4 + len(sorted_entries)}"

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # SHEET 2: Partition Summary — Max Duration DESC
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    ws2 = wb.create_sheet("Partition Summary")
    ws2.sheet_properties.tabColor = "4472C4"

    ws2.merge_cells("A1:I1")
    ws2["A1"] = "Partition Summary — Sorted by Highest Single Roll Duration"
    ws2["A1"].font = TITLE_FONT
    ws2.row_dimensions[1].height = 30

    part_stats = defaultdict(lambda: {"durations": [], "offsets": []})
    for e in entries:
        part_stats[e["partition"]]["durations"].append(e["duration_ms"])
        part_stats[e["partition"]]["offsets"].append(e["offset"])

    p_headers = [
        ("Rank", 7),
        ("Partition Name", 65),
        ("Roll Count", 11),
        ("Max (ms)", 12),
        ("Min (ms)", 12),
        ("Avg (ms)", 12),
        ("Median (ms)", 12),
        ("Total (ms)", 14),
        ("Status", 12),
    ]
    add_headers(ws2, 3, p_headers)

    sorted_parts = sorted(part_stats.items(), key=lambda x: max(x[1]["durations"]), reverse=True)

    for i, (part, stats) in enumerate(sorted_parts):
        row = i + 4
        d = sorted(stats["durations"])
        cnt = len(d)
        max_d = max(d)

        ws2.cell(row=row, column=1, value=i + 1).font = Font(name="Arial", size=10, bold=True)
        ws2.cell(row=row, column=1).alignment = Alignment(horizontal="center")
        ws2.cell(row=row, column=2, value=part).font = DATA_FONT
        ws2.cell(row=row, column=3, value=cnt).font = NUM_FONT
        ws2.cell(row=row, column=3).alignment = Alignment(horizontal="center")
        max_cell = ws2.cell(row=row, column=4, value=max_d)
        style_duration(max_cell, max_d)
        ws2.cell(row=row, column=5, value=min(d)).font = NUM_FONT
        ws2.cell(row=row, column=5).number_format = "#,##0"
        ws2.cell(row=row, column=6, value=round(sum(d) / cnt, 1)).font = NUM_FONT
        ws2.cell(row=row, column=6).number_format = "#,##0.0"
        ws2.cell(row=row, column=7, value=d[cnt // 2]).font = NUM_FONT
        ws2.cell(row=row, column=7).number_format = "#,##0"
        ws2.cell(row=row, column=8, value=sum(d)).font = NUM_FONT
        ws2.cell(row=row, column=8).number_format = "#,##0"
        style_status(ws2.cell(row=row, column=9), max_d)

        if i % 2 == 1:
            for col in [1, 2, 3, 5, 6, 7, 8]:
                ws2.cell(row=row, column=col).fill = ALT_FILL
        for col in range(1, 10):
            ws2.cell(row=row, column=col).border = BORDER

    ws2.freeze_panes = "A4"
    ws2.auto_filter.ref = f"A3:I{3 + len(sorted_parts)}"

    # Bar chart: Top 15 slowest partitions
    if sorted_parts:
        chart = BarChart()
        chart.type = "bar"
        chart.title = "Top Slowest Partitions (Max Roll Duration ms)"
        chart.x_axis.title = "Duration (ms)"
        chart.style = 10
        chart.width = 28
        chart.height = 16
        top_n = min(15, len(sorted_parts))
        data_ref = Reference(ws2, min_col=4, min_row=3, max_row=3 + top_n)
        cats_ref = Reference(ws2, min_col=2, min_row=4, max_row=3 + top_n)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.shape = 4
        ws2.add_chart(chart, f"A{4 + len(sorted_parts) + 2}")

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # SHEET 3: Dashboard
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    ws3 = wb.create_sheet("Dashboard")
    ws3.sheet_properties.tabColor = "7030A0"
    wb.move_sheet(ws3, offset=-2)

    ws3.merge_cells("A1:D1")
    ws3["A1"] = "Kafka Log Segment Roll — Dashboard"
    ws3["A1"].font = Font(name="Arial", bold=True, size=16, color="1F3864")
    ws3.row_dimensions[1].height = 35
    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 55

    if entries:
        all_d = sorted(e["duration_ms"] for e in entries)
        p95 = all_d[int(len(all_d) * 0.95)]
        p99 = all_d[min(int(len(all_d) * 0.99), len(all_d) - 1)]
        slowest = sorted_entries[0]

        stats = [
            ("Total Roll Events", len(entries)),
            ("Unique Partitions", len(part_stats)),
            ("Time Range", f"{min(e['timestamp'] for e in entries).strftime('%Y-%m-%d %H:%M:%S')} to {max(e['timestamp'] for e in entries).strftime('%Y-%m-%d %H:%M:%S')}"),
            ("", ""),
            ("── Duration Stats ──", ""),
            ("Minimum", f"{min(all_d)} ms"),
            ("Maximum", f"{max(all_d)} ms"),
            ("Average", f"{round(sum(all_d)/len(all_d), 1)} ms"),
            ("Median", f"{all_d[len(all_d)//2]} ms"),
            ("P95", f"{p95} ms"),
            ("P99", f"{p99} ms"),
            ("Total Roll Time", f"{sum(all_d):,} ms  ({round(sum(all_d)/1000, 2)} sec)"),
            ("", ""),
            ("── Slowest Partition ──", ""),
            ("Partition", slowest["partition"]),
            ("Duration", f"{slowest['duration_ms']:,} ms ({round(slowest['duration_ms']/1000, 2)} sec)"),
            ("Offset", f"{slowest['offset']:,}"),
            ("When", slowest["timestamp"].strftime("%Y-%m-%d %H:%M:%S")),
            ("", ""),
            ("── Status Breakdown ──", ""),
            ("OK  (< 100 ms)", sum(1 for d in all_d if d < 100)),
            ("MODERATE  (100-999 ms)", sum(1 for d in all_d if 100 <= d < 1000)),
            ("SLOW  (>= 1000 ms)", sum(1 for d in all_d if d >= 1000)),
        ]

        for i, (label, value) in enumerate(stats):
            row = i + 3
            if not label:
                continue
            lc = ws3.cell(row=row, column=1, value=label)
            vc = ws3.cell(row=row, column=2, value=value)
            if "──" in label:
                lc.font = Font(name="Arial", bold=True, size=11, color="2F5496")
                ws3.merge_cells(f"A{row}:B{row}")
            else:
                lc.font = Font(name="Arial", bold=True, size=10)
                vc.font = Font(name="Arial", size=10)
                lc.fill = SUMMARY_FILL
                vc.fill = SUMMARY_FILL
                lc.border = BORDER
                vc.border = BORDER
            if "SLOW" in str(label):
                vc.font = WARN_FONT
                vc.fill = WARN_FILL

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # SHEET 4: Duration Distribution
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    ws4 = wb.create_sheet("Duration Distribution")
    ws4.sheet_properties.tabColor = "70AD47"

    ws4.merge_cells("A1:D1")
    ws4["A1"] = "Roll Duration Distribution"
    ws4["A1"].font = TITLE_FONT
    ws4.row_dimensions[1].height = 30

    buckets = [
        ("0-2 ms", 0, 2), ("3-5 ms", 3, 5), ("6-10 ms", 6, 10),
        ("11-50 ms", 11, 50), ("51-100 ms", 51, 100), ("101-500 ms", 101, 500),
        ("501-1000 ms", 501, 1000), ("1001-5000 ms", 1001, 5000), ("5000+ ms", 5001, float('inf')),
    ]
    all_dur = [e["duration_ms"] for e in entries]

    d_headers = [("Duration Range", 18), ("Count", 10), ("Percentage", 12), ("Visual", 40)]
    for col_idx, (h, w) in enumerate(d_headers, 1):
        cell = ws4.cell(row=3, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = PatternFill("solid", fgColor="548235")
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER
        ws4.column_dimensions[get_column_letter(col_idx)].width = w

    bc = [(lbl, sum(1 for d in all_dur if lo <= d <= hi)) for lbl, lo, hi in buckets]
    mx = max(c for _, c in bc) if bc else 1

    for i, (label, count) in enumerate(bc):
        row = i + 4
        pct = count / len(entries) * 100 if entries else 0
        ws4.cell(row=row, column=1, value=label).font = DATA_FONT
        ws4.cell(row=row, column=2, value=count).font = NUM_FONT
        ws4.cell(row=row, column=2).alignment = Alignment(horizontal="center")
        ws4.cell(row=row, column=3, value=round(pct, 1)).font = NUM_FONT
        ws4.cell(row=row, column=3).number_format = '0.0"%"'
        ws4.cell(row=row, column=4, value="█" * int(count / mx * 30)).font = Font(name="Arial", size=10, color="548235")
        for col in range(1, 5):
            ws4.cell(row=row, column=col).border = BORDER
        if i % 2 == 1:
            for col in range(1, 5):
                ws4.cell(row=row, column=col).fill = PatternFill("solid", fgColor="E2EFDA")

    if bc:
        chart = BarChart()
        chart.type = "col"
        chart.title = "Log Segment Roll Duration Distribution"
        chart.y_axis.title = "Count"
        chart.style = 10
        chart.width = 22
        chart.height = 14
        chart.add_data(Reference(ws4, min_col=2, min_row=3, max_row=3 + len(bc)), titles_from_data=True)
        chart.set_categories(Reference(ws4, min_col=1, min_row=4, max_row=3 + len(bc)))
        ws4.add_chart(chart, "A15")

    wb.save(output_path)
    print(f"\n✅ Report saved: {output_path}")
    print(f"   📊 {len(entries)} entries | {len(part_stats)} unique partitions")


if __name__ == "__main__":
    log_path = sys.argv[1] if len(sys.argv) > 1 else "server.log"
    output = sys.argv[2] if len(sys.argv) > 2 else "kafka_log_segment_report.xlsx"
    if not os.path.exists(log_path):
        print(f"❌ File not found: {log_path}")
        sys.exit(1)
    print(f"🔍 Parsing: {log_path}")
    entries = parse_log(log_path)
    if not entries:
        print("⚠  No 'Rolled new log segment' entries found.")
        sys.exit(1)
    print(f"   Found {len(entries)} entries")
    create_excel(entries, output)
